import openpyxl
import re
import datetime

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

class TikTokParser:
    def __init__(self, log_func):
        self.log = log_func


    def parse_tiktok_excel(self, local_path):
            """Parse file Excel doanh thu TikTok → parsed_json để gửi lên server"""
            if not HAS_OPENPYXL:
                self.log("⚠️ Không có openpyxl, bỏ qua parse TikTok Excel")
                return None
            try:
                wb = openpyxl.load_workbook(local_path, data_only=True)
    
                # ── Sheet Reports: tổng hợp tháng ────────────────────────
                ws_reports = wb["Reports"] if "Reports" in wb.sheetnames else wb.active
                rows = list(ws_reports.iter_rows(values_only=True))
    
                def find_exact(keyword):
                    for r in rows:
                        for c in range(1, 5):  # cột index 1-4
                            if r[c] and str(r[c]).strip() == keyword:
                                return float(r[5] or 0)
                    return 0.0
    
                def ab(v): return abs(float(v or 0))
    
                total_settlement   = find_exact("Total settlement amount")
                total_revenue      = find_exact("Total Revenue")
                subtotal_after     = find_exact("Subtotal after seller discounts")
                refund_subtotal    = ab(find_exact("Refund subtotal after seller discounts"))
                actual_shipping    = find_exact("Actual shipping fee")
                platform_ship_disc = find_exact("Platform shipping fee discount")
                customer_ship_fee  = find_exact("Customer shipping fee")
                actual_return_ship = find_exact("Actual return shipping fee")
                net_shipping_cost  = abs(actual_shipping + platform_ship_disc + customer_ship_fee + actual_return_ship)
                transaction_fee    = ab(find_exact("Transaction fee"))
                commission_fee     = ab(find_exact("TikTok Shop commission fee"))
                order_handling_fee = ab(find_exact("Order processing fee"))
                sfr_service_fee    = ab(find_exact("SFR service fee"))
                flash_sale_fee     = ab(find_exact("Flash Sale service fee"))
                affiliate_fee      = ab(find_exact("Affiliate Commission"))
                affiliate_ads_fee  = ab(find_exact("Affiliate Shop Ads commission"))
                total_affiliate    = affiliate_fee + affiliate_ads_fee
                tax_vat            = ab(find_exact("VAT withheld by TikTok Shop"))
                tax_pit            = ab(find_exact("PIT withheld by TikTok Shop"))
                gmv_tiktok_ads     = ab(find_exact("GMV Payment for TikTok Ads"))
                total_adjustments  = find_exact("Total adjustments")
                fee_total = transaction_fee + commission_fee + order_handling_fee + sfr_service_fee + flash_sale_fee + total_affiliate
                tax_total = tax_vat + tax_pit
    
                _month = ""
                for r in rows:
                    if r[1] and str(r[1]).strip() == "Time period:":
                        import re
                        mp = re.search(r'(\d{4})/(\d{2})', str(r[5] or ""))
                        if mp:
                            _month = f"{mp.group(1)}-{mp.group(2)}"
                        break
    
                # ── Sheet Order details: phí từng đơn ────────────────────
                order_details = []
                if "Order details" in wb.sheetnames:
                    ws_detail = wb["Order details"]
                    detail_headers = []
                    for i, row in enumerate(ws_detail.iter_rows(values_only=True)):
                        if i == 0:
                            detail_headers = [str(c).strip() if c else "" for c in row]
                            continue
                        if not any(row):
                            continue
                        r = dict(zip(detail_headers, row))
                        order_id = str(r.get("Order/adjustment ID  ", r.get("Order/adjustment ID", "")) or "").strip()
                        row_type = str(r.get("Type ", r.get("Type", "")) or "").strip()
                        if row_type != "Order" or len(order_id) < 5:
                            continue
    
                        def g(key): return abs(float(r.get(key, 0) or 0))
    
                        order_details.append({
                            "order_id":       order_id,
                            "fee_commission": g("TikTok Shop commission fee"),
                            "fee_payment":    g("Transaction fee"),
                            "fee_service":    g("Order processing fee") + g("SFR service fee"),
                            "fee_affiliate":  g("Affiliate Commission"),
                            "fee_piship":     g("Actual shipping fee"),
                            "fee_handling":   0,
                            "fee_ads":        g("GMV Payment for TikTok Ads"),
                            "tax_vat":        g("VAT withheld by TikTok Shop"),
                            "tax_pit":        g("PIT withheld by TikTok Shop"),
                            "total_fees":     g("Total Fees"),
                            "settlement":     float(r.get("Total settlement amount", 0) or 0),
                        })
    
                wb.close()
                self.log(f"📊 Parse TikTok Excel xong: {len(order_details)} đơn có phí thực")
    
                return {
                    "_month": _month,
                    "order_details": order_details,
                    "gross_revenue":       total_revenue,
                    "refund_amount":       refund_subtotal,
                    "net_product_revenue": subtotal_after - refund_subtotal,
                    "platform_subsidy":    0,
                    "seller_voucher":      0,
                    "co_funded_voucher":   0,
                    "shipping_net":        -net_shipping_cost,
                    "fee_commission":      commission_fee,
                    "fee_payment":         transaction_fee,
                    "fee_service":         sfr_service_fee + flash_sale_fee,
                    "fee_affiliate":       total_affiliate,
                    "fee_piship_sfr":      sfr_service_fee,
                    "fee_handling":        order_handling_fee,
                    "fee_ads":             gmv_tiktok_ads,
                    "fee_total":           fee_total,
                    "compensation":        max(0, total_adjustments),
                    "tax_vat":             tax_vat,
                    "tax_pit":             tax_pit,
                    "tax_total":           tax_total,
                    "total_payout":        total_settlement,
                }
            except Exception as e:
                self.log(f"⚠️ Lỗi parse TikTok Excel: {str(e)}")
                return None
    def parse_tiktok_order_excel_local(self, local_path, shop_name):
        """Parse file Excel đơn hàng TikTok (sheet OrderSKUList) → JSON {orders, items}"""
        if not HAS_OPENPYXL:
            self.log("⚠️ Không có openpyxl, bỏ qua parse TikTok Order Excel")
            return None
        try:
            wb = openpyxl.load_workbook(local_path, data_only=True)
            ws = wb["OrderSKUList"] if "OrderSKUList" in wb.sheetnames else wb.active
            headers = []
            orders_map = {}
            items = []
            import datetime

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip() if c else "" for c in row]
                    self.log(f"🔍 [DEBUG TÊN CỘT TIKTOK]: {', '.join(headers)}")
                    continue
                if not any(row):
                    continue
                r = dict(zip(headers, row))
                order_id = str(r.get("Order ID", "") or "").strip()
                if not order_id or order_id == "Platform unique order ID.":
                    continue

                status      = str(r.get("Order Status", "") or "").strip().lower()
                cancel_type = str(r.get("Cancelation/Return Type", "") or "").strip().lower()
                cancel_reason = str(r.get("Cancel Reason", "") or "").strip()
                sku          = str(r.get("Seller SKU", "") or "").strip()
                product_name = str(r.get("Product Name", "") or "").strip()

                try: qty = int(r.get("Quantity") or 1)
                except: qty = 1

                try: revenue_line = float(r.get("SKU Subtotal After Discount") or 0)
                except: revenue_line = 0.0

                try: order_amount = float(r.get("Order Amount") or 0)
                except: order_amount = 0.0

                # Phân loại đơn & Chuẩn hóa shipping_status theo rule mới của shop
                order_type = "normal"
                shipping_status = status or ""
                cancel_reason_lower = cancel_reason.lower()

                if "đã hoàn tất" in status or "completed" in status:
                    if not cancel_type:
                        order_type = "normal"
                        shipping_status = "Đã giao"
                    elif "return" in cancel_type or "refund" in cancel_type:
                        order_type = "return"
                        shipping_status = "Hoàn hàng"
                    else:
                        order_type = "normal"
                        shipping_status = "Đã giao"
                        
                elif "đã hủy" in status or "cancel" in status:
                    if "cancel" in cancel_type and "giao gói hàng thất bại" in cancel_reason_lower:
                        # Đơn giao thất bại mất phí -> Xếp vào nhóm Return để tính phí
                        order_type = "return" 
                        shipping_status = "Giao thất bại"
                    else:
                        # Các lý do hủy khác -> Nhóm Cancel thuần túy
                        order_type = "cancel"
                        shipping_status = "Đã hủy"
                        
                elif "return" in cancel_type or "trả hàng" in status or ("hoàn" in status and "hoàn thành" not in status and "hoàn tất" not in status):
                    order_type = "return"
                    shipping_status = "Hoàn hàng"
                    
                else:
                    # Các trạng thái vận chuyển khác
                    if "awaiting collection" in status or "chờ lấy" in status:
                        shipping_status = "Chờ lấy hàng"
                    elif "in transit" in status or "đang giao" in status:
                        shipping_status = "Đang giao"
                    elif "delivered" in status or "đã giao" in status:
                        shipping_status = "Đã giao"

                # Ngày: ưu tiên tìm mọi dạng tên cột ngày tháng để chống lỗi file tháng/ngày
                raw_date = ""
                for k, v in r.items():
                    if k:
                        kl = str(k).lower()
                        if "paid time" in kl or "created time" in kl or "order time" in kl or "thời gian tạo" in kl or "thời gian thanh toán" in kl:
                            raw_date = str(v).strip()
                            if raw_date: break

                order_date = ""
                if raw_date:
                    date_part = raw_date.split(" ")[0]
                    if "-" in date_part:
                        parts = date_part.split("-")
                        if len(parts) == 3:
                            order_date = f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}" if len(parts[0]) == 4 else f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                    elif "/" in date_part:
                        parts = date_part.split("/")
                        if len(parts) == 3:
                            order_date = f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}" if len(parts[0]) == 4 else f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"

                revenue = order_amount if order_type == "normal" else 0

                if sku:
                    items.append({
                        "order_id": order_id, "sku": sku,
                        "product_name": product_name, "qty": qty,
                        "revenue_line": revenue_line, "cost_real": 0, "cost_invoice": 0,
                    })

                if order_id not in orders_map:
                    orders_map[order_id] = {
                        "order_id": order_id, "platform": "tiktok", "shop": shop_name,
                        "order_date": order_date, "order_type": order_type,
                        "revenue": revenue, "raw_revenue": order_amount,
                        "cancel_reason": cancel_reason, "return_fee": 0,
                        "shipped": 0, "cost_invoice": 0, "cost_real": 0,
                        "fee": 0, "profit_invoice": 0, "profit_real": 0,
                        "tax_flat": 0, "tax_income": 0,
                        "fee_platform": 0, "fee_payment": 0, "fee_affiliate": 0,
                        "fee_ads": 0, "fee_piship": 0, "fee_service": 0,
                        "fee_packaging": 0, "fee_operation": 0, "fee_labor": 0,
                        "discount_shop": 0, "discount_shopee": 0,
                        "discount_combo": 0, "shipping_return_fee": 0,
                        "shipping_status": shipping_status,
                    }

            wb.close()
            self.log(f"📊 Parse TikTok Order Excel xong: {len(orders_map)} đơn, {len(items)} items")
            return {"orders": list(orders_map.values()), "items": items}
        except Exception as e:
            self.log(f"⚠️ Lỗi parse TikTok Order Excel: {str(e)}")
            return None
